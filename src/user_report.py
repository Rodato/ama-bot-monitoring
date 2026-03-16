"""
user_report.py
Genera un Excel con una pestaña por ciudad, listando cada usuario activo
en el rango de fechas con: Nombre, Colegio, Salón, Sesiones Usadas, # Sesiones.

Uso:
    python3 src/user_report.py --from 2026-03-02 --to 2026-03-16
    python3 src/user_report.py --from 2026-03-02 --to 2026-03-16 --out data/reports/
"""

import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import db

# ── Estilos ───────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
HEADER_FONT = Font(color="FFFFFF", bold=True)

COLS       = ["Nombre", "Colegio", "Salón", "Sesiones Usadas", "# Sesiones"]
COL_WIDTHS = [30, 35, 20, 30, 14]


# ── Queries ───────────────────────────────────────────────────────────────────

def _get_cities_in_range(date_from: str, date_to: str) -> list:
    """Ciudades que tuvieron actividad en el rango de fechas."""
    sql = """
        SELECT DISTINCT u.city
        FROM ama.ama_user_info_table_v1 u
        WHERE u.city IS NOT NULL AND u.city <> ''
          AND u.client_number IN (
              SELECT DISTINCT client_number
              FROM ama.ama_session_start_table
              WHERE DATE(created_at AT TIME ZONE 'America/Bogota')
                    BETWEEN :date_from AND :date_to
                AND client_number ~ '^(57|59)'
          )
        ORDER BY u.city
    """
    rows = db.query_df(sql, {"date_from": date_from, "date_to": date_to})
    return rows["city"].tolist()


def _get_users_by_city(city: str, date_from: str, date_to: str) -> pd.DataFrame:
    """Un registro por usuario con sus sesiones usadas en el rango."""
    sql = """
        WITH user_sessions AS (
            SELECT
                s.client_number,
                STRING_AGG('S' || s.session::int::text, ' · '
                           ORDER BY s.session::int) AS sesiones_usadas,
                COUNT(DISTINCT s.session::int)      AS num_sesiones
            FROM ama.ama_session_start_table s
            WHERE DATE(s.created_at AT TIME ZONE 'America/Bogota')
                  BETWEEN :date_from AND :date_to
              AND s.client_number ~ '^(57|59)'
            GROUP BY s.client_number
        ),
        latest_user AS (
            SELECT DISTINCT ON (u.client_number)
                u.client_number, u.name, u.course, u.school
            FROM ama.ama_user_info_table_v1 u
            INNER JOIN user_sessions us USING (client_number)
            WHERE u.city = :city
            ORDER BY u.client_number, u.created_at DESC
        )
        SELECT
            COALESCE(lu.name,   'Sin registrar') AS "Nombre",
            COALESCE(lu.school, 'Sin registrar') AS "Colegio",
            COALESCE(lu.course, 'Sin registrar') AS "Salón",
            us.sesiones_usadas                   AS "Sesiones Usadas",
            us.num_sesiones                      AS "# Sesiones"
        FROM latest_user lu
        INNER JOIN user_sessions us USING (client_number)
        ORDER BY lu.school, lu.course, lu.name
    """
    return db.query_df(sql, {"date_from": date_from, "date_to": date_to, "city": city})


# ── Escritura Excel ───────────────────────────────────────────────────────────

def _write_sheet(ws, df: pd.DataFrame, city: str, date_from: str, date_to: str):
    # Título
    ws.append([f"Usuarios Activos — {city}  |  {date_from} al {date_to}"])
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=13, color="1F4E79")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))

    if df.empty:
        ws.append(["Sin usuarios activos en este período."])
        ws.column_dimensions["A"].width = 50
        return

    # Encabezado
    ws.append(COLS)
    for col_idx in range(1, len(COLS) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for row_idx, row in enumerate(df.itertuples(index=False), start=3):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(
                horizontal="center" if col_idx >= 4 else "left"
            )
            if row_idx % 2 == 1:
                cell.fill = ALT_FILL
        ws.cell(row=row_idx, column=5).font = Font(bold=True)

    # Anchos de columnas
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ── Entry point ───────────────────────────────────────────────────────────────

def _build_workbook(date_from: str, date_to: str) -> Workbook:
    cities = _get_cities_in_range(date_from, date_to)

    wb = Workbook()
    wb.remove(wb.active)

    if not cities:
        ws = wb.create_sheet(title="Sin datos")
        ws.append(["Sin usuarios activos en el período seleccionado."])
        return wb

    for city in cities:
        df = _get_users_by_city(city, date_from, date_to)
        ws = wb.create_sheet(title=city[:31])
        _write_sheet(ws, df, city, date_from, date_to)

    return wb


def generate_user_report_bytes(date_from: str, date_to: str) -> bytes:
    """Genera el Excel en memoria y retorna los bytes (para descarga en Streamlit)."""
    import io
    wb = _build_workbook(date_from, date_to)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_user_report(date_from: str, date_to: str, out_dir: str = "data/reports") -> str:
    os.makedirs(out_dir, exist_ok=True)

    wb = _build_workbook(date_from, date_to)
    out_path = os.path.join(out_dir, f"reporte_usuarios_{date_from}_{date_to}.xlsx")
    wb.save(out_path)
    print(f"[user_report] Excel guardado en: {out_path}")
    return out_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Reporte de usuarios AMA por ciudad")
    parser.add_argument("--from", dest="date_from", required=True, help="Fecha inicio YYYY-MM-DD")
    parser.add_argument("--to",   dest="date_to",   required=True, help="Fecha fin YYYY-MM-DD")
    parser.add_argument("--out",  dest="out_dir", default="data/reports", help="Carpeta de salida")
    args = parser.parse_args()

    generate_user_report(args.date_from, args.date_to, args.out_dir)
