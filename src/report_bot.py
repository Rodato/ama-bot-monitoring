"""
report_bot.py
Genera un Excel con 3 hojas de uso del bot AMA para un rango de fechas.

Uso:
    python3 src/report_bot.py --from 2026-03-02 --to 2026-03-09
    python3 src/report_bot.py --from 2026-03-02 --to 2026-03-09 --out data/reports/
"""

import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

import argparse
from datetime import date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import db

# ── Colores ───────────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL      = PatternFill("solid", fgColor="D6E4F0")
TOTAL_FILL    = PatternFill("solid", fgColor="BDD7EE")
HEADER_FONT   = Font(color="FFFFFF", bold=True)


# ── Pivot helper ──────────────────────────────────────────────────────────────

def pivot_dimension(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convierte (fecha, dimension, gender, usuarios) →
    tabla pivot: fecha | dimension | <gender1> | <gender2> | ... | Total
    """
    if df.empty:
        return df

    pivot = df.pivot_table(
        index=["fecha", "dimension"],
        columns="gender",
        values="usuarios",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()

    pivot.columns.name = None
    pivot["Total"] = pivot.iloc[:, 2:].sum(axis=1)
    pivot["fecha"] = pivot["fecha"].astype(str)
    return pivot


# ── Escritura Excel ───────────────────────────────────────────────────────────

def write_sheet(ws, df: pd.DataFrame, sheet_title: str, dim_label: str):
    """Escribe un DataFrame pivotado en una hoja con formato."""
    if df.empty:
        ws.append(["Sin datos para el período seleccionado."])
        return

    cols = df.columns.tolist()
    # Renombrar columnas de display
    display_cols = []
    for c in cols:
        if c == "fecha":
            display_cols.append("Fecha")
        elif c == "dimension":
            display_cols.append(dim_label)
        else:
            display_cols.append(c)

    # Título
    ws.append([sheet_title])
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=13, color="1F4E79")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))

    # Encabezado
    ws.append(display_cols)
    for col_idx, _ in enumerate(display_cols, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for row_idx, row in enumerate(df.itertuples(index=False), start=3):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left")
            if row_idx % 2 == 1:
                cell.fill = ALT_FILL
        # Total en negrita
        ws.cell(row=row_idx, column=len(cols)).font = Font(bold=True)

    # Ancho de columnas
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    for i in range(3, len(cols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14


# ── Hoja de instrucciones ─────────────────────────────────────────────────────

def write_instructions(wb: Workbook, fecha_inicio: str, fecha_fin: str):
    ws = wb.create_sheet(title="Cómo leer este reporte")

    TITLE_FONT   = Font(bold=True, size=14, color="1F4E79")
    SECTION_FONT = Font(bold=True, size=11, color="2E75B6")
    BODY_FONT    = Font(size=10)
    LABEL_FONT   = Font(bold=True, size=10)

    def add_row(text="", font=None, indent=0):
        ws.append([" " * indent + text])
        if font:
            ws.cell(row=ws.max_row, column=1).font = font

    # Título
    ws.append([f"Reporte de Uso — Bot AMA  |  {fecha_inicio} al {fecha_fin}"])
    ws.cell(row=ws.max_row, column=1).font = TITLE_FONT
    ws.append([])

    # Qué es este reporte
    add_row("¿Qué es este reporte?", SECTION_FONT)
    add_row("Muestra cuántos estudiantes usaron el bot de WhatsApp del programa AMA", BODY_FONT, 2)
    add_row("cada día, desglosados por género y agrupados por tres dimensiones:", BODY_FONT, 2)
    add_row("ciudad, colegio y salón.", BODY_FONT, 2)
    ws.append([])

    # Hojas
    add_row("Hojas del archivo", SECTION_FONT)
    hojas = [
        ("Por Ciudad",  "Usuarios activos por ciudad de residencia del estudiante."),
        ("Por Colegio", "Usuarios activos por institución educativa."),
        ("Por Salón",   "Usuarios activos por salón o curso."),
    ]
    for nombre, desc in hojas:
        ws.append(["  • " + nombre])
        ws.cell(row=ws.max_row, column=1).font = LABEL_FONT
        add_row(desc, BODY_FONT, 6)
    ws.append([])

    # Cómo leer las tablas
    add_row("Cómo leer cada tabla", SECTION_FONT)
    columnas = [
        ("Fecha",          "Día calendario en que hubo actividad en el bot."),
        ("Ciudad / Colegio / Salón",
                           "Agrupación. 'Sin registrar' aparece cuando el usuario no completó su perfil."),
        ("Columnas de género", "Cada columna muestra el número de usuarios únicos de ese género "
                               "que enviaron al menos un mensaje ese día. Los valores son "
                               "los que los propios estudiantes registraron al inscribirse."),
        ("Total",          "Suma de todos los géneros para esa fila (fecha + agrupación)."),
    ]
    for col, desc in columnas:
        ws.append(["  • " + col])
        ws.cell(row=ws.max_row, column=1).font = LABEL_FONT
        add_row(desc, BODY_FONT, 6)
    ws.append([])

    # Notas importantes
    add_row("Notas importantes", SECTION_FONT)
    notas = [
        "Un usuario se cuenta UNA SOLA VEZ por día, aunque haya enviado múltiples mensajes.",
        "Los días sin filas = sin actividad registrada en el bot para ese período.",
        "La fuente de datos es la tabla de inicios de sesión del bot (WhatsApp → Supabase).",
        "Los datos demográficos (género, ciudad, colegio, salón) vienen del perfil de registro.",
    ]
    for nota in notas:
        add_row("  — " + nota, BODY_FONT)
    ws.append([])

    # Período
    add_row("Período cubierto", SECTION_FONT)
    add_row(f"  Del {fecha_inicio} al {fecha_fin} (inclusive).", BODY_FONT)

    ws.column_dimensions["A"].width = 90
    ws.sheet_view.showGridLines = False


# ── Resumen para el agente ────────────────────────────────────────────────────

def build_summary(dfs: dict) -> str:
    """Genera un resumen en texto de los 3 DataFrames para pasarle al LLM."""
    lines = []
    for sheet_name, (df, dim_label) in dfs.items():
        lines.append(f"\n### {sheet_name} ({dim_label})")
        if df.empty:
            lines.append("Sin datos.")
            continue
        lines.append(df.to_string(index=False))
    return "\n".join(lines)


# ── Entry point ───────────────────────────────────────────────────────────────

def generate_report(fecha_inicio: str, fecha_fin: str, out_dir: str = "data/reports") -> tuple:
    """
    Genera el Excel y retorna (ruta_excel, resumen_texto).
    """
    os.makedirs(out_dir, exist_ok=True)

    # Dimensiones: (col_db, label_hoja, label_columna)
    dimensions = [
        ("city",   "Por Ciudad",   "Ciudad"),
        ("school", "Por Colegio",  "Colegio"),
        ("course", "Por Salón",    "Salón / Curso"),
    ]

    wb = Workbook()
    wb.remove(wb.active)  # eliminar hoja vacía por defecto

    # Hoja de instrucciones (primera)
    write_instructions(wb, fecha_inicio, fecha_fin)

    dfs = {}
    for col, sheet_name, dim_label in dimensions:
        raw = db.get_usage_by_dimension(col, fecha_inicio, fecha_fin)
        pivoted = pivot_dimension(raw)
        ws = wb.create_sheet(title=sheet_name)
        title = f"Uso del Bot AMA — {sheet_name} | {fecha_inicio} al {fecha_fin}"
        write_sheet(ws, pivoted, title, dim_label)
        dfs[sheet_name] = (pivoted, dim_label)

    out_path = os.path.join(out_dir, f"reporte_bot_{fecha_inicio}_{fecha_fin}.xlsx")
    wb.save(out_path)
    print(f"[report_bot] Excel guardado en: {out_path}")

    summary = build_summary(dfs)
    return out_path, summary


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--from", dest="fecha_inicio", required=True, help="Fecha inicio YYYY-MM-DD")
    parser.add_argument("--to",   dest="fecha_fin",    required=True, help="Fecha fin YYYY-MM-DD")
    parser.add_argument("--out",  dest="out_dir", default="data/reports", help="Carpeta de salida")
    args = parser.parse_args()

    path, _ = generate_report(args.fecha_inicio, args.fecha_fin, args.out_dir)
    print(f"Reporte generado: {path}")
